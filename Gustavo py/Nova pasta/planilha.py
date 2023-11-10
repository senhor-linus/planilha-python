import openpyxl

#criar planilha
book = openpyxl.Workbook()

#como visualizar paginas existentes
print(book.sheetnames)

#como criar uma pagina 
book.create_sheet("computadores")
#como selecionar uma pagina
computadores_page = book['computadores']
computadores_page.append(['LOJAS', 'MEMÓRIA RAM', 'PREÇO'])
computadores_page.append(['Computadores1' , '8 gb ram', 'R$:2.500'])
computadores_page.append(['computadores2', '16 gb ram',  'R$5.500'])
computadores_page.append(['computadores3', '32 gb ram',  'R$8.500'])
# salvar planilha
book.save('Planilha de Conpras.xlsx')